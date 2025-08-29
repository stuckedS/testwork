import sys
import os
import datetime
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtCore import Qt
from openpyxl import load_workbook, Workbook


def is_numb(value):
    try:
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        s = str(value).strip().replace(" ", "")
        if s == "":
            return False
        float(s)
        return True
    except:
        return False


def to_numb(value):
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(" ", "")
        return float(s)
    except:
        return 0.0


def parse_date(value):
    try:
        if value is None:
            return None
        if isinstance(value, datetime.datetime):
            return value
        s = str(value).strip()
        if s == "":
            return None
        return datetime.datetime.strptime(s, "%d.%m.%Y")
    except:
        return None


def firts_low(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    parts = s.split()
    return parts[0].lower() if parts else ""


class startWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Выбор Excel")
        self.setGeometry(300, 300, 420, 150)
        self.selected_file = None
        self.file_label = QtWidgets.QLabel("Файл Excel:")
        self.file_path = QtWidgets.QLineEdit()
        self.file_path.setReadOnly(True)
        self.file_btn = QtWidgets.QPushButton("Выбрать файл")
        self.next_btn = QtWidgets.QPushButton("Далее")
        layout = QtWidgets.QGridLayout()
        layout.addWidget(self.file_label, 0, 0)
        layout.addWidget(self.file_path, 0, 1)
        layout.addWidget(self.file_btn, 0, 2)
        layout.addWidget(self.next_btn, 1, 0, 1, 3)
        self.setLayout(layout)
        self.file_btn.clicked.connect(self.choose_file)
        self.next_btn.clicked.connect(self.open_file)

    def choose_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Выбрать Excel", "", "Excel Files (*.xlsx)")
        if path:
            self.selected_file = path
            self.file_path.setText(path)

    def open_file(self):
        if not self.selected_file:
            QtWidgets.QMessageBox.warning(self, "Ошибка", "Выберите файл!")
            return
        self.close()
        self.parser = excelParser(self.selected_file, parent_start=self)
        self.parser.show()


class excelParser(QtWidgets.QWidget):
    def __init__(self, file_path, parent_start=None):
        super().__init__()
        self.parent_start = parent_start
        self.file_path = file_path
        self.data = []
        self.headers = []
        self.filtered = []
        self.header_row_index = None
        self.col_types = {}
        self.init_ui()
        self.load_file()

    def init_ui(self):
        self.setWindowTitle("excelParser")
        self.setGeometry(200, 200, 1000, 640)

        self.column_label = QtWidgets.QLabel("Столбец:")
        self.column_select = QtWidgets.QComboBox()
        self.value_label = QtWidgets.QLabel("Фильтр:")
        self.value_input = QtWidgets.QLineEdit()

        self.load_btn = QtWidgets.QPushButton("Загрузить новый файл")
        self.close_btn = QtWidgets.QPushButton("Закрыть")
        self.save_btn = QtWidgets.QPushButton("Сохранить результат")
        self.status_label = QtWidgets.QLabel("")

        self.table = QtWidgets.QTableWidget()
        self.table.setSortingEnabled(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        self.table.setWordWrap(True)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("QTableWidget {font-size: 14px;}")

        top_controls = QtWidgets.QHBoxLayout()
        top_controls.addWidget(self.column_label)
        top_controls.addWidget(self.column_select)
        top_controls.addWidget(self.value_label)
        top_controls.addWidget(self.value_input)
        top_controls.addWidget(self.save_btn)
        top_controls.addWidget(self.load_btn)

        main_layout = QtWidgets.QVBoxLayout()
        main_layout.addLayout(top_controls)
        main_layout.addWidget(self.table)
        main_layout.addWidget(self.status_label)
        main_layout.addSpacing(5)
        
        buttons_layout = QtWidgets.QHBoxLayout()
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.close_btn)
        main_layout.addLayout(buttons_layout)

        self.setLayout(main_layout)

        self.value_input.textChanged.connect(self.apply_filter)
        self.column_select.currentIndexChanged.connect(self.apply_filter)
        self.save_btn.clicked.connect(self.save_result)
        self.load_btn.clicked.connect(self.load_new_file)
        self.close_btn.clicked.connect(self.close_application)

        corner_button = self.table.findChild(QtWidgets.QAbstractButton)
        if corner_button:
            corner_button.clicked.connect(self.clear_filt)
        self.table.horizontalHeader().sectionClicked.connect(self.sort_column)

    def clear_filt(self):
        self.value_input.clear()
        self.apply_filter()

    def load_new_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Выбрать Excel", "", "Excel Files (*.xlsx)")
        if not path:
            return
        self.file_path = path
        self.load_file()

    def close_application(self):
        QtWidgets.QApplication.quit()

    def find_head_row(self, rows):
        required = {"ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"}
        for idx, row in enumerate(rows):
            values = [str(c).strip() if c is not None else "" for c in row]
            if any(r in values for r in required):
                return idx, values
        return None, None

    def find_col_types(self, rows):
        types = {}
        for c in range(len(self.headers)):
            col_vals = [row[c] for row in rows if c < len(row)]
            if any(parse_date(v) is not None for v in col_vals):
                types[c] = "date"
            elif any(is_numb(v) for v in col_vals):
                types[c] = "numeric"
            else:
                types[c] = "text"
        self.col_types = types

    def load_file(self):
        try:
            rows = []
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active
            rows = [r for r in ws.iter_rows(values_only=True) if any((c not in (None, "")) for c in r)]
            self.header_row_index, headers = self.find_head_row(rows)
            if not headers:
                self.status_label.setText("Не удалось найти заголовки")
                return
            self.headers = headers
            self.data = rows[self.header_row_index + 1:]
            self.data = [r for r in self.data if any(c not in (None, "") for c in r)]
            self.column_select.clear()
            self.column_select.addItems(self.headers)
            self.find_col_types(self.data)
            self.apply_filter()
            self.status_label.setText(f"Файл: {os.path.basename(self.file_path)} (заголовки в строке {self.header_row_index + 1})")
        except Exception as e:
            self.status_label.setText(f"Ошибка загрузки: {e}")

    def firts_low(self, value):
        if value is None:
            return ""
        s = str(value).strip()
        if s == "":
            return ""
        parts = s.split()
        return parts[0].lower() if parts else ""

    def apply_filter(self):
        if not self.headers or not self.data:
            return
        col_name = self.column_select.currentText()
        filter_val = self.value_input.text().strip().lower()
        col_idx = self.headers.index(col_name) if col_name in self.headers else 0
        if filter_val == "":
            self.filtered = [r for r in self.data if any(c not in (None, "") for c in r)]
        else:
            self.filtered = []
            for row in self.data:
                if col_idx >= len(row):
                    continue
                token = self.firts_low(row[col_idx])
                if token.startswith(filter_val):
                    self.filtered.append(row)
        self.show_preview(self.filtered)

    def show_preview(self, rows):
        self.table.clear()
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c in range(len(self.headers)):
                val = row[c] if c < len(row) else None
                item = QtWidgets.QTableWidgetItem()
                if val is None or val == "":
                    item.setText("")
                    item.setData(Qt.UserRole, "")
                else:
                    typ = self.col_types.get(c, "text")
                    if typ == "date":
                        dt = parse_date(val)
                        if dt:
                            item.setData(Qt.UserRole, dt)
                            item.setText(dt.strftime("%d.%m.%Y"))
                        else:
                            item.setData(Qt.UserRole, "")
                            item.setText(str(val))
                    elif typ == "numeric":
                        num = to_numb(val)
                        item.setData(Qt.UserRole, num)
                        if abs(num - int(num)) < 1e-9:
                            item.setText(str(int(num)))
                        else:
                            item.setText(str(num))
                    else:
                        item.setData(Qt.UserRole, str(val).lower())
                        item.setText(str(val))
                self.table.setItem(r, c, item)
        header = self.table.horizontalHeader()
        for i in range(self.table.columnCount()):
            header.setSectionResizeMode(i, QtWidgets.QHeaderView.Stretch)
        self.table.resizeRowsToContents()

    def sort_column(self, index):
        if not self.filtered:
            return
        col_type = self.col_types.get(index, "text")
        reverse = self.table.horizontalHeader().sortIndicatorOrder() == Qt.DescendingOrder
        if col_type == "date":
            def key_fn(r):
                v = r[index] if index < len(r) else None
                dt = parse_date(v)
                return dt or datetime.datetime.min
        elif col_type == "numeric":
            def key_fn(r):
                v = r[index] if index < len(r) else None
                return to_numb(v)
        else:
            def key_fn(r):
                v = r[index] if index < len(r) else ""
                return str(v).lower()
        try:
            self.filtered.sort(key=key_fn, reverse=reverse)
        except Exception:
            self.filtered.sort(key=lambda r: str(r[index]) if index < len(r) else "", reverse=reverse)
        self.show_preview(self.filtered)

    def resizeEvent(self, event):
        h = max(10, min(18, int(self.height() / 50)))
        self.table.setStyleSheet(f"QTableWidget {{font-size: {h}px;}}")
        super().resizeEvent(event)

    def closeEvent(self, event):
        if self.parent_start:
            self.parent_start.show()
        event.accept()

    def save_result(self):
        if not self.filtered:
            self.status_label.setText("Нет данных для сохранения")
            return
        required = ["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]
        selected = [self.headers.index(c) for c in required if c in self.headers]
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить Excel", "", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            wb = Workbook()
            ws = wb.active
            ws.append([c for c in required])
            for row in self.filtered:
                ws.append([row[i] if i < len(row) else "" for i in selected])
            wb.save(path)
            self.status_label.setText(f"Файл сохранён: {path}")
        except Exception as e:
            self.status_label.setText(f"Ошибка: {e}")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    start = startWindow()
    start.show()
    sys.exit(app.exec_())
